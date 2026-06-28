#!/usr/bin/env python3
"""
Parse the Gymreapers H1 2026 line plan workbooks + SKU master into the JSON
files consumed by /line-plan. Run once when the workbooks are refreshed by
DAT; Phase 1 ingests are manual (no upload endpoint yet).

Input files (place under apparel-analytics/data/line-plan-source/):
  - GR_H1_2026_Line_Plan_2027SS.xlsx
  - GR_H1_2026_Line_Plan_ByChannel.xlsx
  - GR_MASTER_SKU_CATALOG.xlsx

Output files (written to apparel-analytics/public/line-plan/):
  - categories.json        Category Summary (4 categories x monthly trend + FP&A var)
  - titles.json            Line Plan tab (~497 titles with full metrics)
  - channel_summary.json   Channel x Category roll-up
  - titles_by_channel.json Line Plan by Channel (~1107 rows)
  - sku_rollup.json        SKU master rolled up by category x sub-category x collection
  - meta.json              Source filenames, parsed-at, totals
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "line-plan-source"
# Output writes to the Pi's webhook-server data dir, NOT public/. This data is
# confidential (FP&A targets + actuals) — it must never enter the static export
# or the GitHub repo. Webhook server gates it behind /pulse/line-plan + auth.
OUT = Path("/mnt/data/agents/webhook-server/data/line-plan")


def clean(v):
    """Round floats to 2dp where they're $ amounts; pass through everything else."""
    if isinstance(v, float):
        if v != v:  # NaN
            return None
        return round(v, 4)
    return v


def rows_with_header(ws, header_row_idx: int):
    """Iterate data rows as dicts keyed by the header row."""
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else f"_col{i}" for i, h in enumerate(rows[header_row_idx])]
    for row in rows[header_row_idx + 1:]:
        if all(v is None for v in row):
            continue
        yield {header[i]: clean(row[i]) for i in range(min(len(header), len(row)))}


def parse_categories(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Category Summary"]
    # Header row is index 2 (rows 1-2 are title/subtitle)
    raw = list(rows_with_header(ws, 2))

    # The sheet has 3 metric rows per category (Net Sales / Units / Var $).
    # Pivot into one record per category with all 3 metrics.
    by_cat: dict[str, dict] = {}
    for r in raw:
        cat = r.get("Category")
        metric = r.get("Metric")
        if not cat or not metric or cat == "GRAND TOTAL":
            continue
        rec = by_cat.setdefault(cat, {
            "category": cat,
            "months": {},
            "h1_actual": None,
            "h1_target": None,
            "var_dollars": None,
            "var_pct": None,
            "title_count": None,
            "color_count": None,
            "avg_dollars_per_title": None,
        })
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun (partial)"]
        if metric == "Net Sales ($)":
            rec["months"]["net_sales"] = {m: r.get(m) for m in months}
            rec["h1_actual"] = r.get("H1 Total")
            rec["h1_target"] = r.get("H1 Target")
            rec["var_dollars"] = r.get("Var $")
            rec["var_pct"] = r.get("Var %")
            rec["title_count"] = r.get("# Titles")
            rec["color_count"] = r.get("# Colors")
            rec["avg_dollars_per_title"] = r.get("Avg $ / Title")
        elif metric == "Units":
            rec["months"]["units"] = {m: r.get(m) for m in months}
        elif metric == "Var $ vs Target":
            rec["months"]["var_dollars"] = {m: r.get(m) for m in months}

    return list(by_cat.values())


def parse_titles(path: Path, sku_master_path: Path | None = None) -> list[dict]:
    """Parse Line Plan tab. If sku_master_path is given, override the workbook's
    per-title # Colors / Color List with the non-bundle color set from the SKU
    master, because the workbook count includes bundle combo strings (e.g. a
    3-pack with 35 'colors' that are really combinations of 5-6 real ones)."""
    # Build title → non-bundle colors lookup from the SKU master.
    nonbundle_colors: dict[str, set[str]] = {}
    if sku_master_path and sku_master_path.exists():
        wb_sku = openpyxl.load_workbook(sku_master_path, data_only=True)
        ws_sku = wb_sku["Master_SKU_Catalog"]
        rows = list(ws_sku.iter_rows(values_only=True))
        header = [str(h).strip() if h is not None else "" for h in rows[2]]
        idx = {h: i for i, h in enumerate(header) if h}
        for row in rows[3:]:
            title = row[idx["PRODUCT_TITLE"]] if "PRODUCT_TITLE" in idx else None
            if not title:
                continue
            is_bundle = row[idx["IS_BUNDLE"]] is True if "IS_BUNDLE" in idx else False
            if is_bundle:
                continue
            color = row[idx["PRODUCT_COLOR"]] if "PRODUCT_COLOR" in idx else None
            if color:
                nonbundle_colors.setdefault(str(title).strip(), set()).add(str(color).strip())

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Line Plan"]
    out = []
    for r in rows_with_header(ws, 2):
        title = r.get("Product Title")
        if not title:
            continue
        title_str = str(title).strip()
        # Prefer the non-bundle color set when we have it; fall back to the
        # workbook's # Colors otherwise (e.g. for any title not in the SKU master).
        nb_colors = nonbundle_colors.get(title_str)
        if nb_colors is not None:
            color_count = len(nb_colors)
            color_list = sorted(nb_colors)
        else:
            color_count = r.get("# Colors")
            color_list = [c.strip() for c in (r.get("Color List") or "").split(",") if c.strip()]
        out.append({
            "category": r.get("Category"),
            "sub_category": r.get("Sub-Category"),
            "title": title_str,
            "h1_net_sales": r.get("H1 Net Sales ($)"),
            "h1_target": r.get("H1 FP&A Target ($)"),
            "var_dollars": r.get("Var $ vs Target"),
            "var_pct": r.get("Var % vs Target"),
            "pct_total_revenue": r.get("% Total Revenue"),
            "pct_category_revenue": r.get("% Cat Revenue"),
            "h1_units": r.get("H1 Units"),
            "asp": r.get("ASP ($)"),
            "discount_rate": r.get("Discount Rate"),
            "return_rate": r.get("Return Rate"),
            "msrp": r.get("MSRP ($)"),
            "abc_rank": r.get("ABC Rank"),
            "status": r.get("Status"),
            "color_count": color_count,
            "color_list": color_list,
            "monthly": {
                "Jan": r.get("Jan NS ($)"),
                "Feb": r.get("Feb NS ($)"),
                "Mar": r.get("Mar NS ($)"),
                "Apr": r.get("Apr NS ($)"),
                "May": r.get("May NS ($)"),
                "Jun": r.get("Jun NS ($)"),
            },
        })
    return out


def parse_channel_summary(path: Path) -> dict:
    """Channel Summary tab has merged headers for channel groupings — read raw cells."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Channel Summary"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 2 has channel group names, row 3 has Actual/Target/Var per channel
    channel_row = rows[2]
    channels = []
    seen = set()
    for i, v in enumerate(channel_row):
        if v and v not in ("Category", "Metric", "Total Actual", "Total Target", "Var %") and v not in seen:
            channels.append((i, str(v)))
            seen.add(v)

    # Build column index for each (channel, metric)
    sub_row = rows[3]
    col_map = {}  # (channel_name, metric) -> col_idx
    current_channel = None
    for i, v in enumerate(sub_row):
        # Find the most recent channel header at or before col i
        for start, name in channels:
            if start <= i:
                current_channel = name
            else:
                break
        if v in ("Actual $", "Target $", "Var %") and current_channel:
            col_map[(current_channel, v)] = i

    # Data rows start at index 4. Net Sales ($) rows only.
    categories = []
    for row in rows[4:]:
        cat = row[0]
        metric = row[1] if len(row) > 1 else None
        if not cat or metric != "Net Sales ($)" or cat == "GRAND TOTAL":
            continue
        cat_data = {"category": cat, "channels": {}}
        for (chan, met), col_idx in col_map.items():
            chan_data = cat_data["channels"].setdefault(chan, {})
            chan_data[met.split()[0].lower()] = clean(row[col_idx]) if col_idx < len(row) else None
        categories.append(cat_data)

    return {
        "channels": [c[1] for c in channels],
        "categories": categories,
    }


def parse_titles_by_channel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Line Plan by Channel"]
    out = []
    for r in rows_with_header(ws, 2):
        if not r.get("Product Title") or not r.get("Channel"):
            continue
        out.append({
            "category": r.get("Category"),
            "sub_category": r.get("Sub-Category"),
            "title": r.get("Product Title"),
            "channel": r.get("Channel"),
            "h1_net_sales": r.get("H1 Net Sales ($)"),
            "h1_target": r.get("H1 FP&A Target ($)"),
            "var_dollars": r.get("Var $ vs Target"),
            "var_pct": r.get("Var % vs Target"),
            "units": r.get("Units"),
            "asp": r.get("ASP ($)"),
            "status": r.get("Status"),
            "msrp": r.get("MSRP ($)"),
            "abc_rank": r.get("ABC Rank"),
        })
    return out


def parse_sku_rollup(path: Path) -> dict:
    """Roll up the 12k+ SKU master to (category, sub-category, collection) buckets.
    For each bucket, capture: # SKUs, # unique titles, # unique colors, price range,
    status mix, top suppliers. Keep payload small enough for static import."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Master_SKU_Catalog"]

    # Header at row index 2 (rows 1-2 are title/blank)
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[2]]
    idx = {h: i for i, h in enumerate(header) if h}

    def g(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    buckets = defaultdict(lambda: {
        "skus": 0,
        "titles": set(),
        "colors": set(),
        "msrps": [],
        "statuses": defaultdict(int),
        "abc_ranks": defaultdict(int),
        "suppliers": defaultdict(int),
    })

    for row in rows[3:]:
        cat = g(row, "GENERAL_CATEGORY")
        sub = g(row, "SUB_CATEGORY")
        if not cat or not sub:
            continue
        coll = g(row, "COLLECTION") or "N/A"
        key = (cat, sub, coll)
        b = buckets[key]
        b["skus"] += 1
        title = g(row, "PRODUCT_TITLE")
        if title:
            b["titles"].add(str(title).strip())
        # Skip bundle SKUs when building the color palette — bundle PRODUCT_COLOR
        # values are combo strings like "Electric Blue / Atlantis / Emerald" that
        # would over-count distinct colors (e.g. one 3-pack title contributed 35).
        # Bundles still count toward sku_count and status_mix.
        is_bundle = g(row, "IS_BUNDLE") is True
        color = g(row, "PRODUCT_COLOR")
        if color and not is_bundle:
            b["colors"].add(str(color).strip())
        msrp = g(row, "MSRP")
        if isinstance(msrp, (int, float)) and msrp > 0:
            b["msrps"].append(float(msrp))
        status = g(row, "ITEM_STATUS")
        if status:
            b["statuses"][str(status)] += 1
        rank = g(row, "ABC_RANK_AGGREGATE")
        if rank:
            b["abc_ranks"][str(rank)] += 1
        sup = g(row, "PRIMARY_SUPPLIER")
        if sup:
            b["suppliers"][str(sup).split(" ")[0]] += 1  # supplier name, first token

    out = []
    for (cat, sub, coll), b in buckets.items():
        msrps = b["msrps"]
        out.append({
            "category": cat,
            "sub_category": sub,
            "collection": coll,
            "sku_count": b["skus"],
            "title_count": len(b["titles"]),
            "color_count": len(b["colors"]),
            # Include the actual color list so downstream rollups can union
            # across collections (rather than sum the per-collection counts,
            # which would double-count any color that appears in multiple
            # collections, e.g. "Black" in Axis + Infinity + Performance).
            "colors": sorted(b["colors"]),
            "msrp_min": round(min(msrps), 2) if msrps else None,
            "msrp_max": round(max(msrps), 2) if msrps else None,
            "msrp_avg": round(sum(msrps) / len(msrps), 2) if msrps else None,
            "status_mix": dict(b["statuses"]),
            "abc_mix": dict(b["abc_ranks"]),
            "top_suppliers": dict(sorted(b["suppliers"].items(), key=lambda kv: -kv[1])[:3]),
        })

    # Sort: largest sub-category buckets first
    out.sort(key=lambda r: -r["sku_count"])
    return {
        "total_skus": sum(r["sku_count"] for r in out),
        "total_buckets": len(out),
        "rollup": out,
    }


def main():
    if not SRC.exists():
        print(f"Source dir {SRC} does not exist. Create it and drop the 3 xlsx files in.", file=sys.stderr)
        sys.exit(1)
    OUT.mkdir(parents=True, exist_ok=True)

    line_plan_xlsx = SRC / "GR_H1_2026_Line_Plan_2027SS.xlsx"
    channel_xlsx = SRC / "GR_H1_2026_Line_Plan_ByChannel.xlsx"
    sku_xlsx = SRC / "GR_MASTER_SKU_CATALOG.xlsx"

    for p in (line_plan_xlsx, channel_xlsx, sku_xlsx):
        if not p.exists():
            print(f"Missing source file: {p}", file=sys.stderr)
            sys.exit(1)

    print(f"Parsing {line_plan_xlsx.name}...")
    categories = parse_categories(line_plan_xlsx)
    titles = parse_titles(line_plan_xlsx, sku_xlsx)

    print(f"Parsing {channel_xlsx.name}...")
    channel_summary = parse_channel_summary(channel_xlsx)
    titles_by_channel = parse_titles_by_channel(channel_xlsx)

    print(f"Parsing {sku_xlsx.name} (12k+ rows, may take a moment)...")
    sku_rollup = parse_sku_rollup(sku_xlsx)

    (OUT / "categories.json").write_text(json.dumps(categories, indent=2))
    (OUT / "titles.json").write_text(json.dumps(titles, indent=2))
    (OUT / "channel_summary.json").write_text(json.dumps(channel_summary, indent=2))
    (OUT / "titles_by_channel.json").write_text(json.dumps(titles_by_channel, indent=2))
    (OUT / "sku_rollup.json").write_text(json.dumps(sku_rollup, indent=2))

    meta = {
        "parsed_at": datetime.now(timezone.utc).isoformat(),
        "sources": {
            "line_plan": line_plan_xlsx.name,
            "by_channel": channel_xlsx.name,
            "sku_master": sku_xlsx.name,
        },
        "totals": {
            "categories": len(categories),
            "titles": len(titles),
            "title_channel_rows": len(titles_by_channel),
            "sku_buckets": sku_rollup["total_buckets"],
            "total_skus": sku_rollup["total_skus"],
        },
    }
    (OUT / "meta.json").write_text(json.dumps(meta, indent=2))

    print("\nDone. Wrote to:", OUT)
    print(json.dumps(meta, indent=2))


if __name__ == "__main__":
    main()
